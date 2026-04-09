"""
Comparativa de entrenamiento de un perceptrón simple con diferentes
funciones de activación, umbrales y tasas de aprendizaje para las
compuertas lógicas AND y OR.

Genera una tabla en consola y exporta los resultados a Excel (.xlsx)
o CSV como fallback.

Escalado adaptativo: si una configuración no converge, se duplican las
épocas (100 → 200 → 400 → …) hasta alcanzar MSE = 0 o el tope de 10 000.
"""

import csv
from itertools import product
from pathlib import Path

import numpy as np

# ═══════════════════════════════════════════════════════════════════════════════
# 1) Datos de entrada y compuertas lógicas
# ═══════════════════════════════════════════════════════════════════════════════
X = np.array([[0, 0], [0, 1], [1, 0], [1, 1]], dtype=float)

COMPUERTAS = {
    "AND": np.array([0, 0, 0, 1], dtype=float),
    "OR":  np.array([0, 1, 1, 1], dtype=float),
}

# Agregar columna de sesgo (bias)
Xb = np.hstack([X, np.ones((X.shape[0], 1))])

# ═══════════════════════════════════════════════════════════════════════════════
# 2) Grids de hiperparámetros
# ═══════════════════════════════════════════════════════════════════════════════
FUNCIONES     = ["step", "sign", "sigmoid", "linear"]
UMBRALES      = [0.0, 0.5]
TASAS_LR      = [0.01, 0.05, 0.1, 0.2, 0.3, 0.5, 0.8, 1.0]  # 8 tasas → 128 combos
EPOCAS_INICIO = 100          # Épocas iniciales
EPOCAS_TOPE   = 10_000       # Tope de seguridad para escalado adaptativo
SEMILLA       = 42
TOLERANCIA    = 0.01         # Para sigmoid/linear: delta mínimo para actualizar

# ═══════════════════════════════════════════════════════════════════════════════
# 3) Funciones de activación
# ═══════════════════════════════════════════════════════════════════════════════
def step(z, umbral):
    return 1.0 if z >= umbral else 0.0

def sign(z, _umbral=None):
    return 1.0 if z >= 0 else -1.0

def sigmoid(z, _umbral=None):
    return 1.0 / (1.0 + np.exp(-np.clip(z, -500, 500)))

def linear(z, _umbral=None):
    return z

_ACTIVACIONES = {"step": step, "sign": sign, "sigmoid": sigmoid, "linear": linear}


def clasificar(yout, funcion, umbral):
    """Convierte la salida de activación a 0/1 para comparar con etiquetas."""
    if funcion == "sign":
        return 1 if yout > 0 else 0
    return 1 if yout >= umbral else 0

# ═══════════════════════════════════════════════════════════════════════════════
# 4) Métricas
# ═══════════════════════════════════════════════════════════════════════════════
def mse(y_true, y_pred):
    return float(np.mean((y_true - y_pred) ** 2))

def r2_score(y_true, y_pred):
    ss_res = float(np.sum((y_true - y_pred) ** 2))
    ss_tot = float(np.sum((y_true - np.mean(y_true)) ** 2))
    if ss_tot == 0:
        return 1.0 if ss_res == 0 else 0.0
    return 1.0 - ss_res / ss_tot

# ═══════════════════════════════════════════════════════════════════════════════
# 5) Función de entrenamiento
# ═══════════════════════════════════════════════════════════════════════════════
def entrenar_perceptron(X_bias, y, funcion, umbral, tasa_lr, epocas, seed):
    """
    Entrena un perceptrón simple y devuelve (mse, r2, epoca_convergencia).
    epoca_convergencia es None si no convergió dentro del límite de épocas.
    """
    rng = np.random.default_rng(seed)
    w = rng.uniform(-0.5, 0.5, size=(X_bias.shape[1],))
    activar = _ACTIVACIONES[funcion]

    # Para sign, las etiquetas de entrenamiento son -1/+1
    if funcion == "sign":
        y_train = np.where(y == 0, -1.0, 1.0)
    else:
        y_train = y.copy()

    epoca_convergencia = None

    for epoca in range(1, epocas + 1):
        errores = 0
        for xi, yi in zip(X_bias, y_train):
            z = float(np.dot(xi, w))
            yout = activar(z, umbral)

            if funcion in ("step", "sign"):
                delta = yi - yout
                necesita_ajuste = delta != 0
            else:  # sigmoid, linear
                delta = yi - yout
                necesita_ajuste = abs(delta) > TOLERANCIA

            if necesita_ajuste:
                w += tasa_lr * delta * xi
                errores += 1

        if errores == 0:
            epoca_convergencia = epoca
            break

    # Evaluación final: predicciones binarias (0/1)
    y_pred = np.array([
        clasificar(activar(float(np.dot(xi, w)), umbral), funcion, umbral)
        for xi in X_bias
    ], dtype=float)

    return mse(y, y_pred), r2_score(y, y_pred), epoca_convergencia


def entrenar_con_escalado(X_bias, y, funcion, umbral, tasa_lr, seed):
    """
    Entrena escalando las épocas: 100 → 200 → 400 → … hasta convergencia
    (MSE=0) o alcanzar EPOCAS_TOPE.  Devuelve (mse, r2, epoca_conv, epocas_limite).
    """
    epocas = EPOCAS_INICIO
    while epocas <= EPOCAS_TOPE:
        val_mse, val_r2, epoca_conv = entrenar_perceptron(
            X_bias, y, funcion, umbral, tasa_lr, epocas, seed
        )
        if val_mse == 0.0:
            return val_mse, val_r2, epoca_conv, epocas
        epocas *= 2
    # Último intento con el tope
    val_mse, val_r2, epoca_conv = entrenar_perceptron(
        X_bias, y, funcion, umbral, tasa_lr, EPOCAS_TOPE, seed
    )
    return val_mse, val_r2, epoca_conv, EPOCAS_TOPE

# ═══════════════════════════════════════════════════════════════════════════════
# 6) Exportación a Excel / CSV
# ═══════════════════════════════════════════════════════════════════════════════
def exportar_resultados(resultados, ruta_xlsx, ruta_csv):
    encabezados = list(resultados[0].keys())

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Comparativa Perceptrón"

        # Estilo de encabezado
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        center = Alignment(horizontal="center", vertical="center")

        for col_idx, h in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        # Filas de datos
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row_idx, fila in enumerate(resultados, 2):
            for col_idx, h in enumerate(encabezados, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=fila[h])
                cell.alignment = center
                cell.border = thin_border

            # Colorear según convergencia
            convergio = fila["Convergió"]
            fill = green_fill if convergio == "Sí" else red_fill
            for col_idx in range(1, len(encabezados) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

        # Ajustar ancho de columnas
        for col_idx, h in enumerate(encabezados, 1):
            max_len = len(str(h))
            for row_idx in range(2, len(resultados) + 2):
                val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, len(val))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 3

        wb.save(ruta_xlsx)
        print(f"\n📊 Archivo Excel generado: {ruta_xlsx}")

    except ImportError:
        with open(ruta_csv, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=encabezados)
            writer.writeheader()
            writer.writerows(resultados)
        print(f"\n📊 openpyxl no disponible. Archivo CSV generado: {ruta_csv}")

# ═══════════════════════════════════════════════════════════════════════════════
# 7) Programa principal
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    resultados = []

    # Barrido de todas las combinaciones
    combinaciones = list(product(
        COMPUERTAS.items(), FUNCIONES, UMBRALES, TASAS_LR
    ))
    total = len(combinaciones)

    print("=" * 95)
    print("  COMPARATIVA DE ENTRENAMIENTO — PERCEPTRÓN SIMPLE (con escalado adaptativo de épocas)")
    print(f"  Compuertas: AND, OR  |  Épocas: {EPOCAS_INICIO}→{EPOCAS_TOPE:,}  |  Combinaciones: {total}")
    print("=" * 95)

    for i, ((nombre_comp, y), funcion, umbral, tasa) in enumerate(combinaciones, 1):
        val_mse, val_r2, epoca_conv, epocas_usadas = entrenar_con_escalado(
            Xb, y, funcion, umbral, tasa, SEMILLA
        )
        resultados.append({
            "Compuerta":     nombre_comp,
            "Función":       funcion,
            "Umbral":        umbral,
            "Tasa LR":       tasa,
            "MSE":           round(val_mse, 6),
            "R²":            round(val_r2, 6),
            "Época Conv.":   epoca_conv if epoca_conv else "—",
            "Épocas Límite": epocas_usadas,
            "Convergió":     "Sí" if val_mse == 0.0 else "No",
        })
        print(f"  [{i:>3}/{total}] {nombre_comp} {funcion:<8} umbral={umbral} lr={tasa}  "
              f"→ MSE={val_mse:.4f}  épocas={epocas_usadas}", flush=True)

    # ── Tabla en consola ──────────────────────────────────────────────────────
    print(f"\n{'Compuerta':<12}{'Función':<10}{'Umbral':>7}{'Tasa LR':>9}"
          f"{'MSE':>10}{'R²':>10}{'Época':>8}{'Ép.Lím':>8}{'Conv.':>8}")
    print("─" * 82)

    for r in resultados:
        epoca_str = str(r["Época Conv."]).rjust(8)
        print(f"{r['Compuerta']:<12}{r['Función']:<10}{r['Umbral']:>7.1f}{r['Tasa LR']:>9.2f}"
              f"{r['MSE']:>10.4f}{r['R²']:>10.4f}{epoca_str}"
              f"{r['Épocas Límite']:>8}{r['Convergió']:>8}")

    # ── Resumen de no convergidos ─────────────────────────────────────────────
    no_conv = [r for r in resultados if r["Convergió"] == "No"]
    if no_conv:
        print(f"\n⚠️  {len(no_conv)} configuración(es) NO convergieron a MSE=0 "
              f"(incluso con hasta {EPOCAS_TOPE:,} épocas):")
        for r in no_conv:
            print(f"   {r['Compuerta']} {r['Función']:<8} umbral={r['Umbral']} "
                  f"lr={r['Tasa LR']}  MSE={r['MSE']}")
    else:
        print(f"\n✅ ¡Todas las {total} configuraciones convergieron a MSE=0!")

    # ── Mejor configuración por compuerta ─────────────────────────────────────
    print("\n" + "=" * 82)
    print("  MEJORES CONFIGURACIONES (MSE=0, menor épocas)")
    print("=" * 82)

    for nombre_comp in COMPUERTAS:
        subset = [r for r in resultados if r["Compuerta"] == nombre_comp]
        mejor = min(subset, key=lambda r: (
            r["MSE"],
            r["Época Conv."] if isinstance(r["Época Conv."], int) else 99999
        ))
        print(f"\n  🏆 {nombre_comp}:")
        print(f"     Función: {mejor['Función']}  |  Umbral: {mejor['Umbral']}"
              f"  |  Tasa LR: {mejor['Tasa LR']}")
        print(f"     MSE: {mejor['MSE']:.6f}  |  R²: {mejor['R²']:.6f}"
              f"  |  Convergió: {mejor['Convergió']} (época {mejor['Época Conv.']})")

    # ── Mejor configuración global ────────────────────────────────────────────
    mejor_global = min(resultados, key=lambda r: (
        r["MSE"],
        r["Época Conv."] if isinstance(r["Época Conv."], int) else 99999
    ))
    print(f"\n  🥇 MEJOR GLOBAL:")
    print(f"     {mejor_global['Compuerta']} — {mejor_global['Función']}"
          f"  |  Umbral: {mejor_global['Umbral']}"
          f"  |  Tasa LR: {mejor_global['Tasa LR']}")
    print(f"     MSE: {mejor_global['MSE']:.6f}  |  R²: {mejor_global['R²']:.6f}"
          f"  |  Convergió: {mejor_global['Convergió']} (época {mejor_global['Época Conv.']})")

    # ── Resumen promedio por función de activación ────────────────────────────
    print(f"\n{'─' * 82}")
    print("  PROMEDIO POR FUNCIÓN DE ACTIVACIÓN")
    print(f"{'─' * 82}")
    print(f"  {'Función':<12}{'MSE prom':>12}{'R² prom':>12}{'% Convergió':>14}")
    print(f"  {'─' * 48}")

    for fn in FUNCIONES:
        subset = [r for r in resultados if r["Función"] == fn]
        avg_mse = np.mean([r["MSE"] for r in subset])
        avg_r2 = np.mean([r["R²"] for r in subset])
        pct_conv = sum(1 for r in subset if r["Convergió"] == "Sí") / len(subset) * 100
        print(f"  {fn:<12}{avg_mse:>12.6f}{avg_r2:>12.6f}{pct_conv:>13.1f}%")

    # ── Exportar ──────────────────────────────────────────────────────────────
    carpeta = Path(__file__).resolve().parent
    exportar_resultados(
        resultados,
        carpeta / "comparativa_perceptron_resultados.xlsx",
        carpeta / "comparativa_perceptron_resultados.csv",
    )


if __name__ == "__main__":
    main()
